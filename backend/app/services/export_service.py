"""Export service for data export functionality"""

import csv
import json
import io
from datetime import datetime
from typing import Optional, List, Generator
from uuid import UUID

from sqlalchemy.orm import Session

from app.models.result import Result
from app.exceptions import ValidationError


class ExportService:
    """Service for exporting results in various formats"""

    SUPPORTED_FORMATS = ["csv", "json", "excel"]
    CHUNK_SIZE = 1000  # Process results in chunks for large datasets

    @staticmethod
    def export_to_csv(
        results: List[Result],
        columns: Optional[List[str]] = None,
        stream: bool = False,
    ) -> str | Generator[str, None, None]:
        """
        Export results to CSV format.
        
        Args:
            results: List of Result objects
            columns: List of columns to include (None = all)
            stream: If True, return generator for streaming large datasets
            
        Returns:
            CSV string or generator of CSV chunks
        """
        if not results:
            return ""

        # Get all unique keys from data
        all_keys = set()
        for result in results:
            if isinstance(result.data, dict):
                all_keys.update(result.data.keys())

        # Determine columns
        if columns:
            data_columns = [col for col in columns if col in all_keys]
        else:
            data_columns = sorted(list(all_keys))

        # Include metadata columns
        metadata_columns = ["id", "task_id", "source_url", "extracted_at", "created_at"]
        all_columns = metadata_columns + data_columns

        if stream:
            return ExportService._stream_csv(results, all_columns, data_columns)
        else:
            return ExportService._generate_csv(results, all_columns, data_columns)

    @staticmethod
    def _generate_csv(
        results: List[Result],
        all_columns: List[str],
        data_columns: List[str],
    ) -> str:
        """Generate complete CSV string"""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=all_columns)
        writer.writeheader()

        for result in results:
            row = {
                "id": str(result.id),
                "task_id": str(result.task_id),
                "source_url": result.source_url,
                "extracted_at": result.extracted_at.isoformat(),
                "created_at": result.created_at.isoformat(),
            }

            # Add data columns
            if isinstance(result.data, dict):
                for col in data_columns:
                    row[col] = result.data.get(col, "")
            else:
                for col in data_columns:
                    row[col] = ""

            writer.writerow(row)

        return output.getvalue()

    @staticmethod
    def _stream_csv(
        results: List[Result],
        all_columns: List[str],
        data_columns: List[str],
    ) -> Generator[str, None, None]:
        """Stream CSV in chunks"""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=all_columns)
        writer.writeheader()
        yield output.getvalue()

        for result in results:
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=all_columns)

            row = {
                "id": str(result.id),
                "task_id": str(result.task_id),
                "source_url": result.source_url,
                "extracted_at": result.extracted_at.isoformat(),
                "created_at": result.created_at.isoformat(),
            }

            # Add data columns
            if isinstance(result.data, dict):
                for col in data_columns:
                    row[col] = result.data.get(col, "")
            else:
                for col in data_columns:
                    row[col] = ""

            writer.writerow(row)
            yield output.getvalue()

    @staticmethod
    def export_to_json(
        results: List[Result],
        columns: Optional[List[str]] = None,
        stream: bool = False,
    ) -> str | Generator[str, None, None]:
        """
        Export results to JSON format.
        
        Args:
            results: List of Result objects
            columns: List of columns to include (None = all)
            stream: If True, return generator for streaming large datasets
            
        Returns:
            JSON string or generator of JSON chunks
        """
        if not results:
            return "[]"

        if stream:
            return ExportService._stream_json(results, columns)
        else:
            return ExportService._generate_json(results, columns)

    @staticmethod
    def _generate_json(
        results: List[Result],
        columns: Optional[List[str]] = None,
    ) -> str:
        """Generate complete JSON string"""
        data = []
        for result in results:
            item = {
                "id": str(result.id),
                "task_id": str(result.task_id),
                "source_url": result.source_url,
                "extracted_at": result.extracted_at.isoformat(),
                "created_at": result.created_at.isoformat(),
                "data": result.data if isinstance(result.data, dict) else {},
            }

            # Filter columns if specified
            if columns:
                filtered_data = {}
                for col in columns:
                    if col in item["data"]:
                        filtered_data[col] = item["data"][col]
                item["data"] = filtered_data

            data.append(item)

        return json.dumps(data, indent=2, default=str)

    @staticmethod
    def _stream_json(
        results: List[Result],
        columns: Optional[List[str]] = None,
    ) -> Generator[str, None, None]:
        """Stream JSON in chunks"""
        yield "[\n"

        for idx, result in enumerate(results):
            item = {
                "id": str(result.id),
                "task_id": str(result.task_id),
                "source_url": result.source_url,
                "extracted_at": result.extracted_at.isoformat(),
                "created_at": result.created_at.isoformat(),
                "data": result.data if isinstance(result.data, dict) else {},
            }

            # Filter columns if specified
            if columns:
                filtered_data = {}
                for col in columns:
                    if col in item["data"]:
                        filtered_data[col] = item["data"][col]
                item["data"] = filtered_data

            json_str = json.dumps(item, default=str)
            if idx < len(results) - 1:
                yield json_str + ",\n"
            else:
                yield json_str + "\n"

        yield "]"

    @staticmethod
    def export_to_excel(
        results: List[Result],
        columns: Optional[List[str]] = None,
    ) -> bytes:
        """
        Export results to Excel format.
        
        Args:
            results: List of Result objects
            columns: List of columns to include (None = all)
            
        Returns:
            Excel file as bytes
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ValidationError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        if not results:
            # Create empty workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # Get all unique keys from data
        all_keys = set()
        for result in results:
            if isinstance(result.data, dict):
                all_keys.update(result.data.keys())

        # Determine columns
        if columns:
            data_columns = [col for col in columns if col in all_keys]
        else:
            data_columns = sorted(list(all_keys))

        # Include metadata columns
        metadata_columns = ["id", "task_id", "source_url", "extracted_at", "created_at"]
        all_columns = metadata_columns + data_columns

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        # Write header
        for col_idx, col_name in enumerate(all_columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = openpyxl.styles.Font(bold=True)

        # Write data
        for row_idx, result in enumerate(results, 2):
            row_data = {
                "id": str(result.id),
                "task_id": str(result.task_id),
                "source_url": result.source_url,
                "extracted_at": result.extracted_at.isoformat(),
                "created_at": result.created_at.isoformat(),
            }

            # Add data columns
            if isinstance(result.data, dict):
                for col in data_columns:
                    row_data[col] = result.data.get(col, "")
            else:
                for col in data_columns:
                    row_data[col] = ""

            for col_idx, col_name in enumerate(all_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_name, "")

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(all_columns, 1):
            max_length = len(col_name)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def get_clipboard_data(
        results: List[Result],
        format_type: str = "text",
    ) -> str:
        """
        Format results for clipboard.
        
        Args:
            results: List of Result objects
            format_type: Format type (text, json, csv, html)
            
        Returns:
            Formatted data as string
        """
        if not results:
            return ""

        if format_type == "json":
            return ExportService._generate_json(results)
        elif format_type == "csv":
            return ExportService._generate_csv(results, [], [])
        elif format_type == "html":
            return ExportService._generate_html(results)
        else:  # text
            return ExportService._generate_text(results)

    @staticmethod
    def _generate_text(results: List[Result]) -> str:
        """Generate plain text format"""
        lines = []
        for result in results:
            lines.append(f"URL: {result.source_url}")
            lines.append(f"Extracted: {result.extracted_at.isoformat()}")
            if isinstance(result.data, dict):
                for key, value in result.data.items():
                    lines.append(f"  {key}: {value}")
            lines.append("")
        return "\n".join(lines)

    @staticmethod
    def _generate_html(results: List[Result]) -> str:
        """Generate HTML table format"""
        if not results:
            return "<table></table>"

        # Get all unique keys from data
        all_keys = set()
        for result in results:
            if isinstance(result.data, dict):
                all_keys.update(result.data.keys())

        data_columns = sorted(list(all_keys))
        metadata_columns = ["source_url", "extracted_at"]
        all_columns = metadata_columns + data_columns

        html = ['<table border="1" cellpadding="5" cellspacing="0">']
        html.append("<thead><tr>")

        for col in all_columns:
            html.append(f"<th>{col}</th>")

        html.append("</tr></thead>")
        html.append("<tbody>")

        for result in results:
            html.append("<tr>")
            html.append(f"<td>{result.source_url}</td>")
            html.append(f"<td>{result.extracted_at.isoformat()}</td>")

            if isinstance(result.data, dict):
                for col in data_columns:
                    value = result.data.get(col, "")
                    html.append(f"<td>{value}</td>")
            else:
                for _ in data_columns:
                    html.append("<td></td>")

            html.append("</tr>")

        html.append("</tbody>")
        html.append("</table>")

        return "\n".join(html)

    @staticmethod
    def validate_export_format(format_type: str) -> None:
        """Validate export format"""
        if format_type not in ExportService.SUPPORTED_FORMATS:
            raise ValidationError(
                f"Unsupported export format: {format_type}. "
                f"Supported formats: {', '.join(ExportService.SUPPORTED_FORMATS)}"
            )
